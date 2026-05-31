from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config.field_groups import EXCEL_RAW_RECORD_FIELD_KEYS
from app.config.field_registry import DATA_TYPE_AMOUNT, DATA_TYPE_INT, get_field_spec
from app.utils.date_utils import settlement_cycle_display_code


class ExcelService:
    """V1.0 Excel 汇总导出服务。"""

    _RAW_HEADER_OVERRIDES = {
        "team_manager_name_snapshot": "团队经理",
        "account_manager_name_snapshot": "客户经理",
        "source_type": "来源",
    }

    def export_company_report(
        self,
        file_path: str,
        company_name: str,
        start_date: str,
        end_date: str,
        dataset: dict,
    ) -> tuple[bool, str]:
        try:
            wb = Workbook()
            wb.remove(wb.active)

            self._write_raw_sheet(wb, company_name, start_date, end_date, dataset.get("raw_records", []))
            self._write_summary_sheet(
                wb,
                "客户经理汇总",
                company_name,
                start_date,
                end_date,
                dataset.get("by_account_manager", []),
            )
            self._write_summary_sheet(
                wb,
                "团队汇总",
                company_name,
                start_date,
                end_date,
                dataset.get("by_team", []),
            )
            self._write_targets_sheet(wb, company_name, start_date, end_date, dataset.get("cycle_targets", []))
            self._write_alert_sheet(wb, company_name, start_date, end_date, dataset.get("alert_rows", []))
            self._write_logs_sheet(wb, company_name, start_date, end_date, dataset.get("import_logs", []))

            target = Path(file_path)
            target.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(target))
            return True, str(target)
        except Exception as exc:  # noqa: BLE001
            return False, f"Excel 导出失败: {exc}"

    @staticmethod
    def _title(company_name: str, start_date: str, end_date: str) -> str:
        return f"{company_name} 汇总（{start_date} ~ {end_date}）"

    def _write_raw_sheet(self, wb: Workbook, company_name: str, start_date: str, end_date: str, rows: list[dict]) -> None:
        field_keys = list(EXCEL_RAW_RECORD_FIELD_KEYS)
        headers = [self._raw_header_for_key(key) for key in field_keys]

        data = []
        for r in rows:
            data.append([self._raw_value_for_key(key, r) for key in field_keys])

        ws = wb.create_sheet("原始日报记录")
        self._write_table(
            ws,
            title=self._title(company_name, start_date, end_date),
            headers=headers,
            rows=data,
            amount_cols=self._raw_columns_by_type(field_keys, DATA_TYPE_AMOUNT),
            int_cols=self._raw_columns_by_type(field_keys, DATA_TYPE_INT),
            pct_cols=set(),
        )

    @classmethod
    def _raw_header_for_key(cls, field_key: str) -> str:
        return cls._RAW_HEADER_OVERRIDES.get(field_key, get_field_spec(field_key).label)

    @staticmethod
    def _raw_value_for_key(field_key: str, row: dict):
        if field_key == "settlement_cycle_code":
            record_date = str(row.get("record_date", ""))
            if record_date.strip():
                return settlement_cycle_display_code(record_date=record_date)
            return settlement_cycle_display_code(cycle_code=str(row.get("settlement_cycle_code", "")))

        spec = get_field_spec(field_key)
        value = row.get(field_key, spec.default)
        if spec.data_type == DATA_TYPE_AMOUNT:
            return float(value or 0)
        if spec.data_type == DATA_TYPE_INT:
            return int(value or 0)
        return value or ""

    @staticmethod
    def _raw_columns_by_type(field_keys: list[str], data_type: str) -> set[int]:
        return {
            idx
            for idx, key in enumerate(field_keys, start=1)
            if get_field_spec(key).data_type == data_type
        }

    def _write_summary_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        company_name: str,
        start_date: str,
        end_date: str,
        rows: list[dict],
    ) -> None:
        headers = [
            "分组",
            "记录数",
            "累计回款金额",
            "累计放款金额",
            "累计邀约",
            "四星客户数",
            "五星客户数",
            "累计签约量",
            "累计优质上门量",
            "签约率",
            "优质上门率",
            "批复率",
            "回款转化率",
            "目标完成进度",
            "结算周期目标",
        ]

        data = []
        for r in rows:
            data.append(
                [
                    r.get("group_name", ""),
                    int(r.get("record_count", 0) or 0),
                    float(r.get("repayment_amount_cumulative", 0) or 0),
                    float(r.get("loan_amount_cumulative", 0) or 0),
                    int(r.get("invitation_cumulative", 0) or 0),
                    int(r.get("four_star_customer_count", 0) or 0),
                    int(r.get("five_star_customer_count", 0) or 0),
                    int(r.get("signing_count_cumulative", 0) or 0),
                    int(r.get("quality_visit_count_cumulative", 0) or 0),
                    r.get("signing_rate"),
                    r.get("quality_visit_rate"),
                    r.get("approval_rate"),
                    r.get("repayment_conversion_rate"),
                    r.get("target_progress"),
                    float(r.get("team_cycle_target", 0) or 0),
                ]
            )

        ws = wb.create_sheet(sheet_name)
        self._write_table(
            ws,
            title=self._title(company_name, start_date, end_date),
            headers=headers,
            rows=data,
            amount_cols={3, 4, 15},
            int_cols={2, 5, 6, 7, 8, 9},
            pct_cols={10, 11, 12, 13, 14},
        )

    def _write_targets_sheet(self, wb: Workbook, company_name: str, start_date: str, end_date: str, rows: list[dict]) -> None:
        headers = ["团队ID", "客户经理ID", "客户经理", "结算周期", "目标金额", "更新时间"]
        data = [
            [
                int(r.get("team_id", 0) or 0),
                int(r.get("account_manager_id", 0) or 0),
                r.get("account_manager_name", ""),
                settlement_cycle_display_code(cycle_code=str(r.get("settlement_cycle_code", ""))),
                float(r.get("target_amount", 0) or 0),
                r.get("updated_at", ""),
            ]
            for r in rows
        ]

        ws = wb.create_sheet("周期目标配置")
        self._write_table(
            ws,
            title=self._title(company_name, start_date, end_date),
            headers=headers,
            rows=data,
            amount_cols={5},
            int_cols={1, 2},
            pct_cols=set(),
        )

    def _write_alert_sheet(self, wb: Workbook, company_name: str, start_date: str, end_date: str, rows: list[dict]) -> None:
        headers = [
            "团队ID",
            "团队",
            "客户经理ID",
            "客户经理",
            "上门目标",
            "上门完成率",
            "上门状态",
            "优质上门目标",
            "优质上门完成率",
            "优质上门状态",
            "回款目标",
            "回款完成率",
            "回款状态",
            "四星客户数预警状态",
            "五星客户数预警状态",
        ]
        data = []
        for row in rows:
            data.append(
                [
                    int(row.get("team_id", 0) or 0),
                    row.get("team_name", ""),
                    int(row.get("account_manager_id", 0) or 0),
                    row.get("account_manager_name", ""),
                    self._int_or_none(row.get("visit_target")),
                    row.get("visit_completion_rate"),
                    row.get("visit_status", ""),
                    self._int_or_none(row.get("quality_visit_target")),
                    row.get("quality_visit_completion_rate"),
                    row.get("quality_visit_status", ""),
                    self._float_or_none(row.get("repayment_target")),
                    row.get("repayment_completion_rate"),
                    row.get("repayment_status", ""),
                    "连续三工作日未达标预警" if row.get("four_star_low_streak_alert") else "",
                    "连续三工作日未达标预警" if row.get("five_star_low_streak_alert") else "",
                ]
            )

        ws = wb.create_sheet("预警明细")
        self._write_table(
            ws,
            title=self._title(company_name, start_date, end_date),
            headers=headers,
            rows=data,
            amount_cols={11},
            int_cols={1, 3, 5, 8},
            pct_cols={6, 9, 12},
        )

    @staticmethod
    def _int_or_none(value):
        if value is None:
            return None
        return int(float(value or 0))

    @staticmethod
    def _float_or_none(value):
        if value is None:
            return None
        return float(value or 0)

    def _write_logs_sheet(self, wb: Workbook, company_name: str, start_date: str, end_date: str, rows: list[dict]) -> None:
        headers = [
            "log_id",
            "import_time",
            "file_name",
            "team_name",
            "settlement_cycle_code",
            "file_path",
            "export_id",
            "template_version",
            "result",
            "message",
            "affected_record_count",
        ]
        data = [
            [
                int(r.get("id", 0) or 0),
                r.get("import_time", ""),
                r.get("file_name", ""),
                r.get("team_name", ""),
                settlement_cycle_display_code(cycle_code=str(r.get("settlement_cycle_code", ""))),
                r.get("file_path", ""),
                r.get("export_id", ""),
                r.get("template_version", ""),
                r.get("result", ""),
                r.get("message", ""),
                int(r.get("affected_record_count", 0) or 0),
            ]
            for r in rows
        ]
        ws = wb.create_sheet("导入日志")
        self._write_table(
            ws,
            title=self._title(company_name, start_date, end_date),
            headers=headers,
            rows=data,
            amount_cols=set(),
            int_cols={1, 9},
            pct_cols=set(),
        )

    def _write_table(
        self,
        ws,
        title: str,
        headers: list[str],
        rows: list[list],
        amount_cols: set[int],
        int_cols: set[int],
        pct_cols: set[int],
    ) -> None:
        ws.cell(row=1, column=1, value=title)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

        for c, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=c, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, row in enumerate(rows, start=3):
            for c_idx, value in enumerate(row, start=1):
                if c_idx in pct_cols and value is None:
                    value = None
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if c_idx in amount_cols:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx in int_cols:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx in pct_cols:
                    if value is None:
                        cell.value = None
                    else:
                        cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        self._auto_width(ws)
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}2"
        ws.freeze_panes = "A3"

    @staticmethod
    def _auto_width(ws) -> None:
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 52)
