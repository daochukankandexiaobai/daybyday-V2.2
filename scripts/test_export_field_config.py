from __future__ import annotations

import shutil
import sys
import tempfile
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))


def _assert(condition, message):
    if not condition:
        raise AssertionError(message)


def _now():
    return datetime.now().isoformat(timespec="seconds")


class _SettingsService:
    def __init__(self):
        self._values = {}

    def get(self, key, default=""):
        return self._values.get(key, default)

    def set(self, key, value):
        self._values[key] = str(value)

    @staticmethod
    def is_strict_template_mode():
        return False


class _TemplateService:
    @staticmethod
    def get_active_template_version():
        return "test-template"


def _insert_master_data(conn):
    now = _now()
    conn.execute(
        """
        INSERT INTO teams (id, region, team_name, team_manager_name, is_active, created_at, updated_at)
        VALUES (1, '测试区域', '测试团队', '测试经理', 1, ?, ?)
        """,
        (now, now),
    )
    conn.execute(
        """
        INSERT INTO account_managers (id, team_id, account_manager_name, is_active, created_at, updated_at)
        VALUES (101, 1, '张三', 1, ?, ?)
        """,
        (now, now),
    )


def _insert_dynamic_field(conn):
    now = _now()
    conn.execute(
        """
        INSERT INTO field_definitions (
            field_key, label, data_type, category, group_key,
            editable, required, default_value, aggregation, formula_id,
            enabled, system_field, storage_type, storage_column,
            created_at, updated_at
        )
        VALUES (
            'dynamic_export_count', '动态导出数', 'int', 'raw_daily', 'process_behavior',
            1, 0, '0', 'sum', '',
            1, 0, 'dynamic_metric', '',
            ?, ?
        )
        """,
        (now, now),
    )
    for page_key, display_order in [("json_export", 500), ("excel_export", 500)]:
        conn.execute(
            """
            INSERT INTO field_page_visibility (
                field_key, page_key, visible, group_key, display_order, created_at, updated_at
            )
            VALUES ('dynamic_export_count', ?, 1, 'process_behavior', ?, ?, ?)
            """,
            (page_key, display_order, now, now),
        )


def _insert_daily_record(record_repo):
    from app.utils.date_utils import settlement_cycle_display_code

    now = _now()
    payload = {
        "record_id": "export-dynamic-001",
        "business_key": "2026-06-01|测试区域|测试团队|张三",
        "record_date": "2026-06-01",
        "region": "测试区域",
        "team_id": 1,
        "team_name_snapshot": "测试团队",
        "team_manager_name_snapshot": "测试经理",
        "account_manager_id": 101,
        "account_manager_name_snapshot": "张三",
        "settlement_cycle_code": settlement_cycle_display_code(record_date="2026-06-01"),
        "repayment_amount_daily": 100.0,
        "visit_count_daily": 3,
        "signing_count_daily": 1,
        "four_star_customer_count_daily": 2,
        "five_star_customer_count_daily": 3,
        "remark": "",
        "version": 1,
        "created_at": now,
        "updated_at": now,
        "template_version": "test-template",
        "record_hash": "hash-export-dynamic-001",
        "source_type": "local",
    }
    return record_repo.insert(payload)


def main():
    from app.db.database import DatabaseManager
    from app.db.repositories import (
        AccountManagerRepository,
        CycleTargetRepository,
        DailyRecordRepository,
        ImportLogRepository,
        TeamRepository,
    )
    from app.fields.field_value_service import FieldValueService
    from app.services.excel_service import ExcelService
    from app.services.export_service import ExportService
    from app.services.import_service import ImportService
    from app.services.record_service import RecordService
    from app.services.team_service import TeamService
    from app.utils.json_utils import load_json_file

    tmp_dir = Path(tempfile.mkdtemp(prefix="daybyday_export_config_test_"))
    try:
        db = DatabaseManager(str(tmp_dir / "export_config.db"))
        db.initialize()
        with db.get_connection() as conn:
            _insert_master_data(conn)
            _insert_dynamic_field(conn)
            conn.commit()

        record_repo = DailyRecordRepository(db)
        team_repo = TeamRepository(db)
        manager_repo = AccountManagerRepository(db)
        cycle_target_repo = CycleTargetRepository(db)
        settings_service = _SettingsService()
        template_service = _TemplateService()
        team_service = TeamService(team_repo, manager_repo, cycle_target_repo, settings_service)
        record_service = RecordService(
            record_repo=record_repo,
            team_repo=team_repo,
            account_manager_repo=manager_repo,
            cycle_target_repo=cycle_target_repo,
            template_service=template_service,
        )
        field_value_service = FieldValueService(db)

        row_id = _insert_daily_record(record_repo)
        field_value_service.set_value(row_id, "dynamic_export_count", 7)
        record = record_repo.get_by_id(row_id)
        _assert(record is not None, "record should exist")

        export_service = ExportService(
            record_service=record_service,
            team_service=team_service,
            settings_service=settings_service,
            template_service=template_service,
        )
        json_field_defs = export_service._json_export_field_definitions()
        json_keys = [row["field_key"] for row in json_field_defs]
        _assert("dynamic_export_count" in json_keys, "json export should include dynamic field")
        exported_record = export_service._record_for_export(record, json_field_defs)
        _assert(exported_record["dynamic_export_count"] == 7, "json export should read dynamic value")

        ok, message, json_path = export_service.export_json(
            mode="某日",
            team_id=1,
            base_date="2026-06-01",
            custom_start="",
            custom_end="",
            output_dir=str(tmp_dir),
        )
        _assert(ok and json_path, "full json export failed: {}".format(message))
        exported_payload = load_json_file(Path(str(json_path)))
        _assert(
            exported_payload["records"][0].get("dynamic_export_count") == 7,
            "full json export should include dynamic field",
        )

        excel_service = ExcelService(db)
        excel_defs = excel_service._raw_field_definitions()
        excel_keys = [row["field_key"] for row in excel_defs]
        _assert("dynamic_export_count" in excel_keys, "excel raw sheet should include dynamic field")
        output_path = tmp_dir / "export_config.xlsx"
        ok, info = excel_service.export_company_report(
            file_path=str(output_path),
            company_name="测试公司",
            start_date="2026-06-01",
            end_date="2026-06-01",
            dataset={
                "raw_records": [record],
                "by_account_manager": [],
                "by_team": [],
                "cycle_targets": [],
                "alert_rows": [],
                "import_logs": [],
            },
        )
        _assert(ok, "excel export failed: {}".format(info))
        wb = load_workbook(str(output_path), data_only=True)
        ws = wb["原始日报记录"]
        headers = [ws.cell(row=2, column=col).value for col in range(1, ws.max_column + 1)]
        _assert("动态导出数" in headers, "excel header should use field label")
        dynamic_col = headers.index("动态导出数") + 1
        _assert(ws.cell(row=3, column=dynamic_col).value == 7, "excel should export dynamic value")

        import_service = ImportService(
            record_repo=record_repo,
            import_log_repo=ImportLogRepository(db),
            settings_service=settings_service,
            template_service=template_service,
            record_service=record_service,
            team_service=team_service,
            team_repo=team_repo,
            account_manager_repo=manager_repo,
        )
        incoming = dict(exported_record)
        incoming["record_id"] = str(uuid.uuid4())
        incoming["record_date"] = "2026-06-02"
        incoming["business_key"] = "2026-06-02|测试区域|测试团队|张三"
        incoming["record_hash"] = ""
        incoming["dynamic_export_count"] = 11
        result, message, affected = import_service._upsert_record(
            raw_record=incoming,
            file_path="dynamic.json",
            file_template_version="test-template",
        )
        _assert(result == "success" and affected == 1, "dynamic json import failed: {}".format(message))
        imported = record_repo.get_by_record_id(incoming["record_id"])
        _assert(imported is not None, "imported record should exist")
        imported_value = field_value_service.get_value(imported, "dynamic_export_count")
        _assert(imported_value == 11, "json import should write dynamic metric value")

        old_format = dict(exported_record)
        old_format.pop("dynamic_export_count", None)
        old_format["record_id"] = str(uuid.uuid4())
        old_format["record_date"] = "2026-06-03"
        old_format["business_key"] = "2026-06-03|测试区域|测试团队|张三"
        old_format["record_hash"] = ""
        result, message, affected = import_service._upsert_record(
            raw_record=old_format,
            file_path="old.json",
            file_template_version="test-template",
        )
        _assert(result == "success" and affected == 1, "old json import should remain compatible: {}".format(message))

        print("[export_config] PASS")
        return 0
    finally:
        shutil.rmtree(str(tmp_dir), ignore_errors=True)


if __name__ == "__main__":
    raise SystemExit(main())
